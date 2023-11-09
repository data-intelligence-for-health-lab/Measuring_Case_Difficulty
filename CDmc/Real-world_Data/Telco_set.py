import os
os.environ["CUDA_VISIBLE_DEVICES"]="-1"   
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

import numpy as np
import pandas as pd
import random

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, OneHotEncoder, StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer

import tensorflow as tf
from tensorflow import keras
from tensorflow.python.keras import utils

from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import multiprocessing
import logging


# Dataset Setting
log_file_name = './Telco.log'
file_name = './Telco.xlsx'
excel_sheet_name = 'telco'
value = "number_of_neuron"
number_of_NNs = 20


# Configure the logger
logging.basicConfig(filename=os.path.join(log_file_name), level=logging.DEBUG)
logging.info('Start of the prediction:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

# Load Data
data = pd.read_csv('./WA_Fn-UseC_-Telco-Customer-Churn.csv') 
df = data

# Drop Customer ID
df = df.drop('customerID',axis=1)
# Convert SeniorCitizen Column to the object
df["SeniorCitizen"] = df["SeniorCitizen"].replace({0:"No",1:"Yes"})
# Convert Churn column to a numeric 0/1 variable
df["Churn"] = df["Churn"].replace({"No":0, "Yes":1})

# Missing data imputation use the same row information
df.loc[df['tenure'] == 0, 'TotalCharges'] = 0
df["TotalCharges"] = df["TotalCharges"].astype('float64')


# Categorical columns and numeric columns
Cat_col = ['gender', 'SeniorCitizen', 'Partner', 'Dependents', 'PhoneService',
       'MultipleLines', 'InternetService', 'OnlineSecurity', 'OnlineBackup',
       'DeviceProtection', 'TechSupport', 'StreamingTV', 'StreamingMovies',
       'Contract', 'PaperlessBilling', 'PaymentMethod']
Num_col = ["tenure", "TotalCharges", "MonthlyCharges"]


# Label Encoder for categorical 
label_encoder = LabelEncoder()
df[Cat_col] = df[Cat_col].apply(label_encoder.fit_transform)


# X and y set
X = df.drop('Churn',axis=1)
y = df['Churn']


# One-hot encoding for categorical 
onehot_encoder = OneHotEncoder(handle_unknown='ignore') 
enc_data = pd.DataFrame(onehot_encoder.fit_transform(X[Cat_col]).toarray())
enc_data.columns = onehot_encoder.get_feature_names()
column_names = enc_data.columns.to_list()

# Merge numeric columns and one-hot encoded columns
X.drop(X[Cat_col] ,axis=1, inplace=True)
X_OH_data = pd.concat([X, enc_data ], axis=1)
X_OH_data

# Scaler Set
scaler_transformer = Pipeline(steps=[('scaler', StandardScaler())])
scaler_preprocessor = ColumnTransformer(transformers=[
    ('scaling', scaler_transformer, Num_col)], remainder='passthrough')


telco_X = X_OH_data
telco_y = y


# reload file
if os.path.isfile(file_name):
    wb = load_workbook(file_name, read_only=True)   # open an Excel file and return a workbook
    if excel_sheet_name in wb.sheetnames:
        logging.info('sheet exist:%s'%excel_sheet_name)
        number_of_index_done = pd.read_excel(file_name,sheet_name=excel_sheet_name,header=None)
        #print(number_of_index_done)

        if len(number_of_index_done) == 0:
            starting = len(number_of_index_done)
            print("Starting index:",starting) #when the excel file is empty

        else:
            if len(number_of_index_done) == len(telco_X):
                print('Done')
                starting = None
            
            else:
                starting = number_of_index_done[0].iloc[-1]+1
                print("Starting index:",starting) #Still need to generate index
    else:
        logging.info("no sheet telco exist, generate")
        wb = load_workbook(file_name)
        wb.create_sheet(excel_sheet_name)
        wb.save(file_name)
        starting = 0

else:
    logging.info("no file exist, generate")
    wb = Workbook()
    ws = wb.active
    ws.title = excel_sheet_name
    wb.save(filename = file_name)
    starting = 0


# manually generate the different random seed
def random_generater():
    return(random.randint(1, 100000))


# Model Complexity (NN)
def nn_model_complexity_multiprocessing(X, y, X_test, y_test,number_of_neuron):
    count = 0 
    X_train, X_val, y_train, y_val = train_test_split(X, y, test_size=0.3, random_state=random_generater())

    X_train_scaled = scaler_preprocessor.fit_transform(X_train)
    X_val_scaled = scaler_preprocessor.transform(X_val)
    X_test_scaled = scaler_preprocessor.transform(X_test)
    
    #print(X_train,"\n") #need to check the train data is shuffled
    #print(X_val,"\n")
    
    model = tf.keras.models.Sequential()
    model.add(tf.keras.layers.Dense(number_of_neuron, input_shape=(X_train_scaled.shape[1],), activation='relu')) #Increasing number of neuron
    model.add(tf.keras.layers.Dense(1, activation='sigmoid')) #output layer
    model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
    
    es = tf.keras.callbacks.EarlyStopping(monitor='val_loss', mode='min', verbose=0, patience=30)
    history = model.fit(X_train_scaled, y_train, validation_data=(X_val_scaled, y_val), batch_size=32, epochs=100, verbose=0, callbacks=[es])
    
    y_pred = (model.predict(X_test_scaled) > 0.5).astype(int)
    #print('y_test:',y_test,' y_pred:',y_pred,"\n")
    
    if y_pred == y_test:
        count += 1
    else:
        count +=0
         
    tf.keras.backend.clear_session()
    output.put(count)
    
rows_value = []    
output = multiprocessing.Queue()
if __name__ == "__main__":
    rows = []
    start_time_overall = datetime.now()
    for index in range(starting,len(telco_X)): #number of index to check
        print("\nindex:",index)
        start_time_one = datetime.now()
        
        X_all = telco_X
        y_all = telco_y
        
        #One sample left out
        X_test = X_all.iloc[[index]] # the test case that want to check the difficulty
        y_test = y_all[index]
    
        X = X_all.drop(index=[index]) # X,y the dataset wilthout the test case
        y = y_all.drop(index=[index])
    
        correct_count = 0
        number_of_neuron = 1 # number of neuron in the hidden layer
        NNs = 0
        
        processes = []
        for NNs in range(number_of_NNs): #How many NN to generate
            p = multiprocessing.Process(target=nn_model_complexity_multiprocessing, args=(X, y, X_test, y_test,number_of_neuron))
            p.start()
            processes.append(p)

        try:
            for process in processes:
                process.join()
                
            # Get process results from the output queue
            correct_count = [output.get() for p in processes]
            print("correct_count:",correct_count)
            
            #print("correct number:",sum(correct_count))
            rows = [index] + [X_test[column][index] for column in X_test.columns] + [y_test]+ [number_of_neuron] + [sum(correct_count)]
            rows_value.append(rows)
            results = pd.DataFrame(rows_value)
            print("results:",results)
            
            end_time_one = datetime.now()
            one_case_time = end_time_one-start_time_one
            logging.info('%s case_running_time:%s',index,one_case_time)
    
            writer = pd.ExcelWriter(file_name)
            results.to_excel(writer, sheet_name=excel_sheet_name, header=None, index=False)
            writer.close()
            
        except KeyboardInterrupt:
            logging.info("parent received ctrl-c")
            for process in processes:
                process.terminate()
                process.join()
        
        
end_time_overall = datetime.now()
overall_case_time = end_time_overall-start_time_overall
logging.info('overall_running_time::%s'%overall_case_time)


# Adding column_names to excel file
new_column_names = ['case_index']+ Num_col + column_names + ['y','number_of_neuron','correct_count']
data_with_column_names = pd.read_excel(file_name,header=None)
data_with_column_names.columns = new_column_names
data_with_column_names.to_excel(file_name, index=False, sheet_name=excel_sheet_name)