import os
os.environ["CUDA_VISIBLE_DEVICES"]="-1"   
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

import numpy as np
import pandas as pd
import random

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer

import tensorflow as tf
from tensorflow import keras
from tensorflow.python.keras import utils

from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import multiprocessing
import logging
import sys


# Dataset Setting
log_file_name = './Telco.log'
file_name = './Telco.xlsx'
excel_sheet_name = 'telco'
value = "number_of_neuron"
number_of_NNs = 20


# Configure the logger
logging.basicConfig(filename=os.path.join(log_file_name), level=logging.DEBUG)
logging.info('Start of the prediction:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


original_data = pd.read_csv('./WA_Fn-UseC_-Telco-Customer-Churn.csv') 
MNN = round(len(original_data)*0.01)
logging.info('MNN:%s'%MNN)


# Scaler Set
Num_col = ["tenure", "TotalCharges", "MonthlyCharges"]
scaler_transformer = Pipeline(steps=[('scaler', StandardScaler())])
scaler_preprocessor = ColumnTransformer(transformers=[
    ('scaling', scaler_transformer, Num_col)], remainder='passthrough')


# reload file
if os.path.isfile(file_name):
    wb = load_workbook(file_name, read_only=True)   # open an Excel file and return a workbook
    if excel_sheet_name in wb.sheetnames:
        logging.info('sheet exist:%s'%excel_sheet_name)
        number_of_index_done = pd.read_excel(file_name,sheet_name=excel_sheet_name)
        #print(number_of_index_done)

        #when the excel file is empty
        if len(number_of_index_done) == 0:
            sys.exit("Error: number of index is not enough.")

        else:
            if len(number_of_index_done) == len(original_data):
                logging.info('Ready to search the MNN')
            
            #Still need to generate index
            else:
                sys.exit("Error: number of index is not enough.")

    #When the sheet not exist
    else:
        logging.info("no sheet exist")
        sys.exit("no sheet exist.")
#When the file not exist
else:
    logging.info("no file exist")
    sys.exit("no file exist.")

    
# manually generate the different random seed
def random_generater():
    return(random.randint(1, 100000))


# Model Complexity (NN)
def nn_model_complexity_multiprocessing(X, y, X_test, y_test,number_of_neuron):
    count = 0 
    X_train, X_val, y_train, y_val = train_test_split(X, y, test_size=0.3, random_state=random_generater())

    X_train_scaled = scaler_preprocessor.fit_transform(X_train)
    X_val_scaled = scaler_preprocessor.transform(X_val)
    X_test_scaled = scaler_preprocessor.transform(X_test)
    
    #print(X_train,"\n") #need to check the train data is shuffled
    #print(X_val,"\n")
    
    model = tf.keras.models.Sequential()
    #Increasing number of neuron
    model.add(tf.keras.layers.Dense(number_of_neuron, input_shape=(X_train_scaled.shape[1],), activation='relu')) 
    #output layer
    model.add(tf.keras.layers.Dense(1, activation='sigmoid'))
    model.compile(loss='binary_crossentropy', optimizer='adam', metrics=['accuracy'])
    
    es = tf.keras.callbacks.EarlyStopping(monitor='val_loss', mode='min', verbose=0, patience=30)
    history = model.fit(X_train_scaled, y_train, validation_data=(X_val_scaled, y_val), batch_size=32, epochs=100, verbose=0, callbacks=[es])
    
    y_pred = (model.predict(X_test_scaled) > 0.5).astype(int)
    #print('y_test:',y_test,' y_pred:',y_pred,"\n")
    
    if y_pred == y_test:
        count += 1
    else:
        count +=0
        
    tf.keras.backend.clear_session()
    output.put(count)
    
    
output = multiprocessing.Queue()
while True:
    start_time_overall = datetime.now()
    updated_rows = []
        
    one_neuron_file = pd.read_excel(file_name, sheet_name=excel_sheet_name)
    X_all = one_neuron_file.drop(columns=['case_index', 'y', 'number_of_neuron', 'correct_count'])
    y_all = one_neuron_file['y'] 
    
    # Check the index that needs to be repeated
    repeat_index_count_view = one_neuron_file.loc[(one_neuron_file["number_of_neuron"] < MNN)&(one_neuron_file["correct_count"] < number_of_NNs*0.9),
                                                  ["case_index","number_of_neuron"]]
    logging.info('The total number of indexes left:%s'%repeat_index_count_view)
    print("The total number of indexes left:",len(repeat_index_count_view))
    print(repeat_index_count_view,"\n")

    
    # Check the index that needs to be repeated with the certain number of neuron
    number_of_neuron = repeat_index_count_view['number_of_neuron'].min()
    repeat_index_neuron_count_view = repeat_index_count_view.loc[(one_neuron_file["number_of_neuron"] <= number_of_neuron),
                                                                 ["case_index","number_of_neuron"]] 
    
    print("The number of indexes left (Neuron perspective):",len(repeat_index_neuron_count_view))
    print(repeat_index_neuron_count_view,"\n")
    
    number_of_neuron += 1
    print("Number of Neuron Used:",number_of_neuron)
    
    if(len(repeat_index_neuron_count_view) == 0):
        logging.info('Done')
        break;

    else:
        for i in repeat_index_neuron_count_view["case_index"].values.tolist():
            start_time_one = datetime.now()
            
            X_test = X_all.iloc[[i]] # the test case that want to check the difficulty
            y_test = y_all[i]
  
            X = X_all.drop(index=[i]) # X,y the dataset wilthout the test case
            y = y_all.drop(index=[i])
    
            processes = []
            for NNs in range(number_of_NNs): #How many NN to generate
                p = multiprocessing.Process(target=nn_model_complexity_multiprocessing, args=(X, y, X_test, y_test,number_of_neuron))
                p.start()
                processes.append(p)

            try:
                for process in processes:
                    process.join()     
                # Get process results from the output queue
                correct_count = [output.get() for p in processes]
                print("index",i, "correct number:",sum(correct_count))

                end_time_one = datetime.now()
                one_case_time = end_time_one-start_time_one
                logging.info('%s case_running_time:%s',i,one_case_time)

                updated_rows.append([i] + [X_test[column][i] for column in X_test.columns] + [y_test]+ [number_of_neuron] + [sum(correct_count)])
                #print(updated_rows)

                score = pd.DataFrame(updated_rows,columns=one_neuron_file.columns)
                result = pd.concat([one_neuron_file, score]).drop_duplicates(['case_index'], keep='last').sort_values('case_index')
                result = result.reset_index(drop=True)
                #print(result)

                writer = pd.ExcelWriter(file_name)
                result.to_excel(writer, sheet_name=excel_sheet_name,index=False)
                writer.close()
            
            except KeyboardInterrupt:
                print("parent received ctrl-c")
                for process in processes:
                    process.terminate()
                    process.join()
                sys.exit("parent received ctrl-c.")
        
        end_time_overall = datetime.now()
        overall_case_time = end_time_overall-start_time_overall
        logging.info('overall_running_time:%s'%overall_case_time)
