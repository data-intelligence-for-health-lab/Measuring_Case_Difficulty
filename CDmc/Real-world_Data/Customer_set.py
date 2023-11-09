import os
os.environ["CUDA_VISIBLE_DEVICES"]= '-1'   
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

import numpy as np
import pandas as pd
import random

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, OneHotEncoder, StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer

import tensorflow as tf
from tensorflow import keras
from keras.utils import np_utils

from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import multiprocessing
import logging


# Dataset Setting
log_file_name = './Customer.log'
file_name = './Customer.xlsx'
excel_sheet_name = 'customer'
value = "number_of_neuron"
number_of_NNs = 20


# Configure the logger
logging.basicConfig(filename=os.path.join(log_file_name), level=logging.DEBUG)
logging.info('Start of the prediction:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

# Customer Segmentation Train data (https://www.kaggle.com/datasets/vetrirah/customer/data)
data = pd.read_csv("Train.csv").drop(columns=["ID"])
df = data

# Preprocessing
# Define the column types
categorical_columns = ['Gender', 'Ever_Married', 'Graduated','Profession','Spending_Score','Var_1']
numeric_columns = ['Age', 'Work_Experience', 'Family_Size']


# categorical columns (imputation, one hot encoder)
# numeric columns (imputation, standardscaler)
categorical_transformer = Pipeline([
    ('imputer', SimpleImputer(strategy='most_frequent')),
    ('one_hot_encoder', OneHotEncoder(handle_unknown='ignore'))
])

numeric_transformer = Pipeline([
    ('imputer', SimpleImputer(strategy='most_frequent')),
    ('scaler', StandardScaler())
])

# Combine the transformers using ColumnTransformer
preprocessor = ColumnTransformer([
    ('categorical', categorical_transformer, categorical_columns),
    ('numeric', numeric_transformer, numeric_columns)], remainder='passthrough')


# X and y set
X = df.drop('Segmentation',axis=1)
y = df['Segmentation']

# changing y (A,B,C,D) to number
y = y.map({'A':0,'B':1,'C':2,'D':3})

# X is before preprocessing
customer_X = X
customer_y = y


# reload file
if os.path.isfile(file_name):
    wb = load_workbook(file_name, read_only=True)   # open an Excel file and return a workbook
    if excel_sheet_name in wb.sheetnames:
        logging.info('sheet exist:%s'%excel_sheet_name)
        number_of_index_done = pd.read_excel(file_name,sheet_name=excel_sheet_name,header=None)
        #print(number_of_index_done)

        if len(number_of_index_done) == 0:
            starting = len(number_of_index_done)
            print("Starting index:",starting) #when the excel file is empty

        else:
            if len(number_of_index_done) == len(customer_X):
                print('Done')
                starting = None
            
            else:
                starting = number_of_index_done[0].iloc[-1]+1
                print("Starting index:",starting) #Still need to generate index
    else:
        logging.info("no sheet exist, generate")
        wb = load_workbook(file_name)
        wb.create_sheet(excel_sheet_name)
        wb.save(file_name)
        starting = 0

else:
    logging.info("no file exist, generate")
    wb = Workbook()
    ws = wb.active
    ws.title = excel_sheet_name
    wb.save(filename = file_name)
    starting = 0
    
    
# manually generate the different random seed
def random_generater():
    return(random.randint(1, 100000))


# Model Complexity (NN)
def nn_model_complexity_multiprocessing(X, y, X_test, y_test,number_of_neuron):
    count = 0 
    
    # multi-class y label one hot encoding
    y = np_utils.to_categorical(y,4)
    y_test = np_utils.to_categorical(y_test,4)  
    
    X_train, X_val, y_train, y_val = train_test_split(X, y, test_size=0.3, random_state=random_generater())

    X_train_scaled = preprocessor.fit_transform(X_train)
    X_val_scaled = preprocessor.transform(X_val)
    X_test_scaled = preprocessor.transform(X_test)
    
    
    #print(X_train,"\n") #need to check the train data is shuffled
    #print(X_val,"\n")
    
    model = tf.keras.models.Sequential()
    model.add(tf.keras.layers.Dense(number_of_neuron, input_shape=(X_train_scaled.shape[1],), activation='relu')) #Increasing number of neuron
    model.add(tf.keras.layers.Dense(4, activation='softmax')) #output layer
    model.compile(loss='categorical_crossentropy', optimizer='adam', metrics=['accuracy'])
    
    es = tf.keras.callbacks.EarlyStopping(monitor='val_loss', mode='min', verbose=0, patience=30)
    history = model.fit(X_train_scaled, y_train, validation_data=(X_val_scaled, y_val), batch_size=32, epochs=100, verbose=0, callbacks=[es])
    
    y_test = np.argmax(y_test)
    y_pred = np.argmax(model.predict(np.array(X_test_scaled)), axis=1)
    
    if np.all(y_pred == y_test):
        count += 1
    else:
        count +=0
         
    tf.keras.backend.clear_session()
    output.put(count)
    
rows_value = []    
output = multiprocessing.Queue()
if __name__ == "__main__":
    rows = []
    start_time_overall = datetime.now()
    for index in range(starting,len(customer_X)): #number of index to check
        print("\nindex:",index)
        start_time_one = datetime.now()
        
        X_all = customer_X
        y_all = customer_y
        
        #One sample left out
        X_test = X_all.iloc[[index]] # the test case that want to check the difficulty
        y_test = y_all[index]
    
        X = X_all.drop(index=[index]) # X,y the dataset wilthout the test case
        y = y_all.drop(index=[index])
    
        correct_count = 0
        number_of_neuron = 1 # number of neuron in the hidden layer
        NNs = 0
        
        processes = []
        for NNs in range(number_of_NNs): #How many NN to generate
            p = multiprocessing.Process(target=nn_model_complexity_multiprocessing, args=(X, y, X_test, y_test,number_of_neuron))
            p.start()
            processes.append(p)

        try:
            for process in processes:
                process.join()
                
            # Get process results from the output queue
            correct_count = [output.get() for p in processes]
            #print("correct_count:",correct_count)
            #print("correct number:",sum(correct_count))
            
            rows = [index] + [X_test[column][index] for column in X_test.columns] + [y_test]+ [number_of_neuron] + [sum(correct_count)]
            print("rows:",rows)
            rows_value.append(rows)
            results = pd.DataFrame(rows_value)
            #print("results:",results)
            
            end_time_one = datetime.now()
            one_case_time = end_time_one-start_time_one
            logging.info('%s case_running_time:%s',index,one_case_time)
    
            writer = pd.ExcelWriter(file_name)
            results.to_excel(writer, sheet_name=excel_sheet_name, header=None, index=False)
            writer.close()
            
        except KeyboardInterrupt:
            logging.info("parent received ctrl-c")
            for process in processes:
                process.terminate()
                process.join()
        
        
end_time_overall = datetime.now()
overall_case_time = end_time_overall-start_time_overall
logging.info('overall_running_time::%s'%overall_case_time)


# Adding column_names to excel file
data_with_column_names = pd.read_excel(file_name,header=None)
if len(data_with_column_names) == len(data):
    data_with_column_names.columns = ['case_index']+ X.columns.to_list() + ['y','number_of_neuron','correct_count']
    data_with_column_names.to_excel(file_name, index=False, sheet_name=excel_sheet_name)
else:
    pass