import os
os.environ['CUDA_VISIBLE_DEVICES'] = '-1'   
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

#scikit-learn k-fold cross-validation
from sklearn.model_selection import KFold
from sklearn.datasets import make_blobs
from sklearn.model_selection import train_test_split

import pandas as pd
import numpy as np
from numpy import array

import random
import itertools

import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
from tensorflow.keras.optimizers import Adam

from keras.callbacks import EarlyStopping
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

from multiprocessing import Pool
import sys
import logging

# Dataset Setting
file_name = './(i)two_class_MS_100.xlsx'
hyper_file_name = './app3_(i)two_ms_data_hyperparam.xlsx'
excel_sheet_name = '(i)'
log_file_name = './(i)two_class_MS_100.log'
number_of_iter = 100
number_of_cpu = 20

# Configure the logger
logging.basicConfig(filename=os.path.join(log_file_name), level=logging.DEBUG)
logging.info('Start of the prediction:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


def moon_shape(number_of_moons,sigma, radius, number_of_datapoints,y_shift):
    moons = []
    np.random.seed(45)
    for y in range(number_of_moons):
        q = np.random.uniform(0,np.pi,size=number_of_datapoints)
        if y % 2 == 0:
            factor = 1
        else: 
            factor = -1  
        moon = np.zeros((number_of_datapoints, 3))
        moon[:,0] = (radius * np.cos(q)) + y
        moon[:,1] = (radius * np.sin(q) * factor) + (factor == -1) * y_shift
        moon[:,2] = y
        moons.append(moon)
        noise = np.random.normal(0, sigma, size=moon[:,:2].shape)
        moon[:,:2] += noise
    moons = np.concatenate(moons)
    logging.info(f"number_of_moons: {number_of_moons}, sigma: {sigma}, number_of_datapoints: {number_of_datapoints}")
    d = {'x1': moons[:,0], 'x2': moons[:,1], 'label': moons[:,2]}
    data = pd.DataFrame(data=d) 
    return(data)


#Two class moon shape
## (g)sigma=0.1 (h)sigma=0.2 (i)sigma=0.4
two_ms_data = moon_shape(number_of_moons=2,sigma=0.4, radius=1, number_of_datapoints=1000,y_shift = 0.3) 
overall_data = two_ms_data


hyper_param_df = pd.read_excel(hyper_file_name, sheet_name=excel_sheet_name, index_col=None, header=None, names=["index","learnRate","batch_size",'activation','hidden_layer_sizes'])  
hyper_param_df['hidden_layer_sizes'] = hyper_param_df['hidden_layer_sizes'].apply(lambda x: eval(x))
if len(hyper_param_df) != len(overall_data):
    sys.exit("Error: Number of hyper_param is not enough.")
    
# reload file
if os.path.isfile(file_name):
    wb = load_workbook(file_name, read_only=True)   # open an Excel file and return a workbook
    if excel_sheet_name in wb.sheetnames:
        logging.info('sheet exist:%s'%excel_sheet_name)
        number_of_index_done = pd.read_excel(file_name,sheet_name=excel_sheet_name,header=None)
        #print(number_of_index_done)

        if len(number_of_index_done) == 0:
            starting = len(number_of_index_done)
            print("Starting index:",starting) #when the excel file is empty

        else:
            if len(number_of_index_done) == len(overall_data):
                print('Done')
                starting = None
            
            else:
                starting = number_of_index_done[0].iloc[-1]+1
                print("Starting index:",starting) #Still need to generate index
                
    else:
        starting = number_of_index_done[0].iloc[-1]+1
        print("Starting index:",starting) #Still need to generate index

else:
    logging.info("no file exist, generate")
    wb = Workbook()
    ws = wb.active
    ws.title = excel_sheet_name
    wb.save(filename = file_name)
    starting = 0
    
# manually generate the different random seed
def random_generater():
    return(random.randint(1, 100000))

# Model Complexity (NN)
def nn_model_complexity_multiprocessing(X_train, X_val, y_train, y_val, X_test,best_params):
    model = tf.keras.models.Sequential()
    # First hidden layer with input shape
    model.add(tf.keras.layers.Dense(best_params['hidden_layer_sizes'][0], input_shape=(2,), activation=best_params['activation']))   
    for i in range(1,len(best_params['hidden_layer_sizes'])):
        # from second hidden layer to number of hidden layers
        model.add(tf.keras.layers.Dense(best_params['hidden_layer_sizes'][i], activation=best_params['activation']))
    # Ouput layer 
    model.add(tf.keras.layers.Dense(1,activation='sigmoid'))
    model.compile(loss='binary_crossentropy', optimizer=Adam(learning_rate=best_params['learnRate']), metrics=['accuracy'])   
    es = EarlyStopping(monitor='val_loss',mode='min', verbose=0, patience=10)     
    model.fit(X_train, y_train, validation_data=(X_val, y_val), verbose=0, 
                       batch_size = best_params['batch_size'], epochs = 100, 
                       callbacks=[es])
    
    difficulty = model.predict(X_test, verbose=0)
    #print(difficulty[0][0])
    return(difficulty[0][0])
    
if __name__ == "__main__":
    row_values = []
    start_time_overall = datetime.now()
    for index in range(starting,len(overall_data)):
        logging.info('running index:%s'%index)
        start_time_one = datetime.now()
        
        X_overall = overall_data[['x1','x2']]
        y_overall = overall_data['label']

        X_test = X_overall.iloc[[index]] # the test case that want to check the difficulty
        y_test = y_overall[index]

        X_without_test = X_overall.drop(index=[index]) # X,y the dataset wilthout the test case
        y_without_test = y_overall.drop(index=[index])
    
        #Call hyperparameter
        param = hyper_param_df.iloc[index]

        X_train_dataset = []
        X_val_dataset = []
        y_train_dataset = []
        y_val_dataset = []
        for nn in range(number_of_iter):
            X_train, X_val, y_train, y_val = train_test_split(X_without_test, y_without_test, test_size=0.3, random_state=random_generater())
            X_train_dataset.append(X_train)
            X_val_dataset.append(X_val)
            y_train_dataset.append(y_train)
            y_val_dataset.append(y_val)
       
    
        predicted_probabilities = []      
        def collect_result(result):
            predicted_probabilities.append(result)
            
        # Create a list of argument tuples
        arg_list = [(X_train_dataset[mm], X_val_dataset[mm], y_train_dataset[mm], y_val_dataset[mm]
                     ,X_test,param) for mm in range(number_of_iter)]
      
        try:
            pool = Pool(processes=number_of_cpu) 
            for args in arg_list:
                pool.apply_async(nn_model_complexity_multiprocessing, args, callback=collect_result)
            pool.close()
            pool.join()
                
        except KeyboardInterrupt:
            print("KeyboardInterrupt detected. Terminating...")
            predicted_probabilities = []
            pool.terminate()
            pool.join()

        end_time_one = datetime.now()
        one_case_time = end_time_one-start_time_one
        logging.info('one case_running_time:%s'%one_case_time)

        if len(predicted_probabilities) != number_of_iter:
            sys.exit("Error: Number of predicted_probabilities does not match with number_of_iter.")
        
        else:
            #index, X_test, y_test, predicted_probabilities(number: number_of_iter)
            row_values = [index] + [X_test[column][index] for column in X_test.columns] + [y_test]+ predicted_probabilities 
            print(row_values)

            wb = load_workbook(filename=file_name)
            ws = wb[excel_sheet_name]
            ws.append(row_values)
            wb.save(file_name)

end_time_overall = datetime.now()
overall_case_time = end_time_overall-start_time_overall
logging.info('overall_running_time:%s'%overall_case_time)
logging.info('End of the prediction_100:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    

