import os
os.environ["CUDA_VISIBLE_DEVICES"]="-1"   
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

# scikit-learn k-fold cross-validation
from numpy import array
from sklearn.model_selection import KFold
import pandas as pd
import numpy as np

import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
from tensorflow.keras.optimizers import Adam

from keras.models import load_model
from sklearn.model_selection import train_test_split

import itertools

from keras.callbacks import EarlyStopping
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook


from ray import tune
from ray.tune.search.hyperopt import HyperOptSearch
import ray
import logging

ray.init(num_cpus=50, num_gpus=0)


# Configure the logger
root_path = './'
logging.basicConfig(filename=os.path.join(root_path, 'app3_(i)two_ms_data_hyperparam.log'), level=logging.DEBUG)
logging.info('Start of the app3:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
file_name = './app3_(i)two_ms_data_hyperparam.xlsx'
excel_sheet_name = '(i)'


def moon_shape(number_of_moons,sigma, radius, number_of_datapoints,y_shift):
    moons = []
    np.random.seed(45)
    for y in range(number_of_moons):
        q = np.random.uniform(0,np.pi,size=number_of_datapoints)
        if y % 2 == 0:
            factor = 1
        else: 
            factor = -1  
        moon = np.zeros((number_of_datapoints, 3))
        moon[:,0] = (radius * np.cos(q)) + y
        moon[:,1] = (radius * np.sin(q) * factor) + (factor == -1) * y_shift
        moon[:,2] = y
        moons.append(moon)
        noise = np.random.normal(0, sigma, size=moon[:,:2].shape)
        moon[:,:2] += noise
    moons = np.concatenate(moons)
    logging.info(f"number_of_moons: {number_of_moons}, sigma: {sigma}, number_of_datapoints: {number_of_datapoints}")
    d = {'x1': moons[:,0], 'x2': moons[:,1], 'label': moons[:,2]}
    data = pd.DataFrame(data=d) 
    return(data)


#Two class moon shape
## (g)sigma=0.1 (h)sigma=0.2 (i)sigma=0.4
two_ms_data = moon_shape(number_of_moons=2,sigma=0.4, radius=1, number_of_datapoints=1000,y_shift = 0.3) 
overall_data = two_ms_data


# reload file
if os.path.isfile(file_name):
    wb = load_workbook(file_name, read_only=True)   # open an Excel file and return a workbook
    if excel_sheet_name in wb.sheetnames:
        logging.info('sheet exist:%s'%excel_sheet_name)
        number_of_index_done = pd.read_excel(file_name,sheet_name=excel_sheet_name,header=None)
        #print(number_of_index_done)

        if len(number_of_index_done) == 0:
            starting = len(number_of_index_done)
            print("Starting index:",starting) #when the excel file is empty

        else:
            if len(number_of_index_done) == len(overall_data):
                print('Done')
                starting = None
            
            else:
                starting = number_of_index_done[0].iloc[-1]+1
                print("Starting index:",starting) #Still need to generate index
                
    else:
        starting = number_of_index_done[0].iloc[-1]+1
        print("Starting index:",starting) #Still need to generate index

else:
    logging.info("no file exist, generate")
    wb = Workbook()
    ws = wb.active
    ws.title = excel_sheet_name
    wb.save(filename = file_name)
    starting = 0
    
    
numbers = [5, 10, 15, 20]
one_combinations_with_order = list(itertools.product(numbers))
two_combinations_with_order = list(itertools.product(numbers, repeat=2))
three_combinations = list(itertools.product(numbers, repeat=3))

layer_neuron_orders = []
layer_neuron_orders = one_combinations_with_order + two_combinations_with_order+three_combinations


def objective(config):
    # Create model
    model = tf.keras.models.Sequential()
    # First hidden layer with input shape
    model.add(Dense(config['hidden_layer_sizes'][0], input_shape=(2,), activation=config['activation']))   
    for i in range(1,len(config['hidden_layer_sizes'])):
        # from second hidden layer to number of hidden layers
        model.add(Dense(config['hidden_layer_sizes'][i], activation=config['activation']))
    model.add(tf.keras.layers.Dense(1,activation='sigmoid'))
    model.compile(loss='binary_crossentropy', optimizer=Adam(learning_rate=config['learnRate']), metrics=['accuracy'])  
    
    es = EarlyStopping(monitor='val_loss',mode='min', verbose=0, patience=30)     
    result = model.fit(X_without_test, y_without_test, validation_split=0.3, verbose=0, 
                       batch_size = config['batch_size'], epochs = 100, 
                       callbacks=[es])
    
    val_loss = result.history['val_loss'][-1]
    return {"val_loss": val_loss}


start_time_overall = datetime.now()
for index in range(starting,len(overall_data)):
    start_time_one = datetime.now()
    X_overall = overall_data[['x1','x2']]
    y_overall = overall_data['label']

    X_test = X_overall.iloc[[index]] # the test case that want to check the difficulty
    y_test = y_overall[index]

    X_without_test = X_overall.drop(index=[index]) # X,y the dataset wilthout the test case
    y_without_test = y_overall.drop(index=[index])

    search_space = {'learnRate': tune.choice([0.01,0.03,0.1]),
                    'batch_size': tune.choice([32,64,128]),
                    'activation':tune.choice(['relu','tanh']),
                    'hidden_layer_sizes':tune.choice(layer_neuron_orders)}

    algo = HyperOptSearch()

    tuner = tune.Tuner(
        objective,
        tune_config=tune.TuneConfig(
            metric="val_loss",
            mode="min",
            num_samples = 100,
            search_alg=algo,
        ),
        run_config = ray.air.config.RunConfig(verbose=0),
        param_space=search_space,
    )

    try:
        results = tuner.fit()
        
        print("Best hyperparameters found were: ", results.get_best_result().config)
        hyper_param = results.get_best_result().config

        rows = [index,hyper_param['learnRate'],hyper_param['batch_size'],
                hyper_param['activation'],hyper_param['hidden_layer_sizes']] 
        logging.info('rows:%s'%rows)
        row_with_string_tuple = rows[:-1] + [str(rows[-1])]

        wb = load_workbook(filename=file_name)
        ws = wb[excel_sheet_name]
        ws.append(row_with_string_tuple)
        wb.save(file_name)

        end_time_one = datetime.now()
        one_case_time = end_time_one-start_time_one
        logging.info('one case_running_time:%s'%one_case_time)


    except KeyboardInterrupt:
        results = None
        ray.shutdown()
        break  

end_time_overall = datetime.now()    
overall_case_time = end_time_overall-start_time_overall
logging.info('overall_running_time:%s'%overall_case_time)
logging.info('End of the app3:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))