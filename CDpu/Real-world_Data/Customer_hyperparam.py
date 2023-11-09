import os
os.environ["CUDA_VISIBLE_DEVICES"]= '-1'   
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

import numpy as np
import pandas as pd
import random

from numpy import array
from sklearn.model_selection import KFold

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, OneHotEncoder, StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer

import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
from tensorflow.keras.optimizers import Adam
from tensorflow import keras
from keras.utils import np_utils

import itertools
import random
from keras.callbacks import EarlyStopping
from datetime import datetime

from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

from ray import tune
from ray.tune.search.hyperopt import HyperOptSearch
import ray
import logging

ray.init(num_cpus=50, num_gpus=0)

# Dataset Setting
log_file_name = './app3_customer_hyperparam.log'
file_name = './app3_customer_hyperparam.xlsx'
excel_sheet_name = 'customer'

# Configure the logger
logging.basicConfig(filename=os.path.join(log_file_name), level=logging.DEBUG)
logging.info('Start of the app3:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


# Load Data & Drop Customer ID
data = pd.read_csv("Train.csv").drop(columns=["ID"])
df = data


# Preprocessing
# Define the column types
categorical_columns = ['Gender', 'Ever_Married', 'Graduated','Profession','Spending_Score','Var_1']
numeric_columns = ['Age', 'Work_Experience', 'Family_Size']

# categorical columns (imputation, one hot encoder)
# numeric columns (imputation, standardscaler)
categorical_transformer = Pipeline([
    ('imputer', SimpleImputer(strategy='most_frequent')),
    ('one_hot_encoder', OneHotEncoder(handle_unknown='ignore'))
])

numeric_transformer = Pipeline([
    ('imputer', SimpleImputer(strategy='most_frequent')),
    ('scaler', StandardScaler())
])

# Combine the transformers using ColumnTransformer
preprocessor = ColumnTransformer([
    ('categorical', categorical_transformer, categorical_columns),
    ('numeric', numeric_transformer, numeric_columns)], remainder='passthrough')


# X and y set
X = df.drop('Segmentation',axis=1)
y = df['Segmentation']

# changing y (A,B,C,D) to number
y = y.map({'A':0,'B':1,'C':2,'D':3})

# X is before preprocessing
customer_X = X
customer_y = y


# reload file
if os.path.isfile(file_name):
    wb = load_workbook(file_name, read_only=True)   # open an Excel file and return a workbook
    if excel_sheet_name in wb.sheetnames:
        logging.info('sheet exist:%s'%excel_sheet_name)
        number_of_index_done = pd.read_excel(file_name,sheet_name=excel_sheet_name,header=None)
        #print(number_of_index_done)

        if len(number_of_index_done) == 0:
            starting = len(number_of_index_done)
            print("Starting index:",starting) #when the excel file is empty

        else:
            if len(number_of_index_done) == len(customer_X):
                print('Done')
                starting = None
            
            else:
                starting = number_of_index_done[0].iloc[-1]+1
                print("Starting index:",starting) #Still need to generate index
    else:
        logging.info("no sheet exist, generate")
        wb = load_workbook(file_name)
        wb.create_sheet(excel_sheet_name)
        wb.save(file_name)
        starting = 0

else:
    logging.info("no file exist, generate")
    wb = Workbook()
    ws = wb.active
    ws.title = excel_sheet_name
    wb.save(filename = file_name)
    starting = 0
    
    
numbers = [5, 10, 15, 20]
one_combinations_with_order = list(itertools.product(numbers))
two_combinations_with_order = list(itertools.product(numbers, repeat=2))
three_combinations = list(itertools.product(numbers, repeat=3))

layer_neuron_orders = []
layer_neuron_orders = one_combinations_with_order + two_combinations_with_order+three_combinations


# manually generate the different random seed
def random_generater():
    return(random.randint(1, 100000))

def objective(config):
    X_train, X_val, y_train, y_val = train_test_split(X_without_test, y_without_test, test_size=0.3, random_state=random_generater())
    
    X_train = preprocessor.fit_transform(X_train)
    X_val = preprocessor.transform(X_val)
    
    # Create model
    model = tf.keras.models.Sequential()
    # First hidden layer with input shape
    model.add(Dense(config['hidden_layer_sizes'][0], input_shape=(X_train.shape[1],), activation=config['activation']))   
    for i in range(1,len(config['hidden_layer_sizes'])):
        # from second hidden layer to number of hidden layers
        model.add(Dense(config['hidden_layer_sizes'][i], activation=config['activation']))
    model.add(tf.keras.layers.Dense(4,activation='softmax'))
    model.compile(loss='categorical_crossentropy', optimizer=Adam(learning_rate=config['learnRate']), metrics=['accuracy'])  
    
    es = EarlyStopping(monitor='val_loss',mode='min', verbose=0, patience=30)     
    result = model.fit(X_train, y_train, validation_data=(X_val, y_val), verbose=0, 
                       batch_size = config['batch_size'], epochs = 100, 
                       callbacks=[es])

    val_loss = result.history['val_loss'][-1]
    return {"val_loss": val_loss}


start_time_overall = datetime.now()
for index in range(starting,len(customer_X)):
    start_time_one = datetime.now()
    X_overall = customer_X
    y_overall = customer_y
    
    X_test = X_overall.iloc[[index]] # the test case that want to check the difficulty
    y_test = y_overall[index]

    X_without_test = X_overall.drop(index=[index]) # X,y the dataset wilthout the test case
    y_without_test = y_overall.drop(index=[index])  
    
    # one-hot encoding for multi-class y
    y_without_test = np_utils.to_categorical(y_without_test,4)

    search_space = {'learnRate': tune.choice([0.01,0.03,0.1]),
                    'batch_size': tune.choice([32,64,128]),
                    'activation':tune.choice(['relu','tanh']),
                    'hidden_layer_sizes':tune.choice(layer_neuron_orders)}

    algo = HyperOptSearch()

    tuner = tune.Tuner(
        objective,
        tune_config=tune.TuneConfig(
            metric="val_loss",
            mode="min",
            num_samples = 100,
            search_alg=algo,
        ),
        run_config = ray.air.config.RunConfig(verbose=0),
        param_space=search_space,
    )

    try:
        results = tuner.fit()
        
        print("Best hyperparameters found were: ", results.get_best_result().config)
        hyper_param = results.get_best_result().config

        rows = [index,hyper_param['learnRate'],hyper_param['batch_size'],
                hyper_param['activation'],hyper_param['hidden_layer_sizes']] 
        logging.info('rows:%s'%rows)
        row_with_string_tuple = rows[:-1] + [str(rows[-1])]

        wb = load_workbook(filename=file_name)
        ws = wb[excel_sheet_name]
        ws.append(row_with_string_tuple)
        wb.save(file_name)

        end_time_one = datetime.now()
        one_case_time = end_time_one-start_time_one
        logging.info('one case_running_time:%s'%one_case_time)


    except KeyboardInterrupt:
        results = None
        ray.shutdown()
        break  

end_time_overall = datetime.now()    
overall_case_time = end_time_overall-start_time_overall
logging.info('overall_running_time:%s'%overall_case_time)
logging.info('End of the app3:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
