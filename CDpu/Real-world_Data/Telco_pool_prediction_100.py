import os
os.environ['CUDA_VISIBLE_DEVICES'] = '-1'   
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

#scikit-learn k-fold cross-validation
from sklearn.model_selection import KFold
from sklearn.model_selection import train_test_split

import pandas as pd
import numpy as np
from numpy import array
from keras.utils import np_utils

import random
import itertools

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, OneHotEncoder, StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer

import logging
logging.getLogger("tensorflow").setLevel(logging.ERROR)


import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
from tensorflow.keras.optimizers import Adam

from keras.callbacks import EarlyStopping
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

from multiprocessing import Pool
import sys

# Dataset Setting
file_name = './Telco_100.xlsx'
hyper_file_name = './app3_telco_hyperparam.xlsx'
excel_sheet_name = '(telco)'
log_file_name = './app3_telco_prediction.log'
number_of_iter = 100
number_of_cpu = 30


# Configure the logger
logging.basicConfig(filename=os.path.join(log_file_name), level=logging.DEBUG)
logging.info('Start of the prediction:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


# Load Data
data = pd.read_csv('./WA_Fn-UseC_-Telco-Customer-Churn.csv') 
df = data

# Drop Customer ID
df = df.drop('customerID',axis=1)
# Convert SeniorCitizen Column to the object
df["SeniorCitizen"] = df["SeniorCitizen"].replace({0:"No",1:"Yes"})
# Convert Churn column to a numeric 0/1 variable
df["Churn"] = df["Churn"].replace({"No":0, "Yes":1})

# Missing data imputation use the same row information
df.loc[df['tenure'] == 0, 'TotalCharges'] = 0
df["TotalCharges"] = df["TotalCharges"].astype('float64')

# Categorical columns and numeric columns
Cat_col = ['gender', 'SeniorCitizen', 'Partner', 'Dependents', 'PhoneService',
       'MultipleLines', 'InternetService', 'OnlineSecurity', 'OnlineBackup',
       'DeviceProtection', 'TechSupport', 'StreamingTV', 'StreamingMovies',
       'Contract', 'PaperlessBilling', 'PaymentMethod']
Num_col = ["tenure", "TotalCharges", "MonthlyCharges"]


# Label Encoder for categorical 
label_encoder = LabelEncoder()
df[Cat_col] = df[Cat_col].apply(label_encoder.fit_transform)


# X and y set
X = df.drop('Churn',axis=1)
y = df['Churn']


# One-hot encoding for categorical 
onehot_encoder = OneHotEncoder(handle_unknown='ignore') 
enc_data = pd.DataFrame(onehot_encoder.fit_transform(X[Cat_col]).toarray())
enc_data.columns = onehot_encoder.get_feature_names_out()
column_names = enc_data.columns.to_list()
names = Num_col + column_names


# Merge numeric columns and one-hot encoded columns
X.drop(X[Cat_col] ,axis=1, inplace=True)
X_OH_data = pd.concat([X, enc_data], axis=1)
X_OH_data


# Scaler Set
scaler_transformer = Pipeline(steps=[('scaler', StandardScaler())])
scaler_preprocessor = ColumnTransformer(transformers=[
    ('scaling', scaler_transformer, Num_col)], remainder='passthrough')


#Split X and y
telco_X = X_OH_data
telco_y = y

hyper_param_df = pd.read_excel(hyper_file_name, sheet_name=excel_sheet_name, index_col=None, header=None, names=["index","learnRate","batch_size",
                                                                               'activation','hidden_layer_sizes'])  
hyper_param_df['hidden_layer_sizes'] = hyper_param_df['hidden_layer_sizes'].apply(lambda x: eval(x))
if len(hyper_param_df) != len(telco_X):
    sys.exit("Error: Number of hyper_param is not enough.")
    
# reload file
if os.path.isfile(file_name):
    wb = load_workbook(file_name, read_only=True)   # open an Excel file and return a workbook
    if excel_sheet_name in wb.sheetnames:
        logging.info('sheet exist:%s'%excel_sheet_name)
        number_of_index_done = pd.read_excel(file_name,sheet_name=excel_sheet_name,header=None)
        #print(number_of_index_done)

        if len(number_of_index_done) == 0:
            starting = len(number_of_index_done)
            print("Starting index:",starting) #when the excel file is empty

        else:
            if len(number_of_index_done) == len(telco_X):
                print('Done')
                starting = None
            
            else:
                starting = number_of_index_done[0].iloc[-1]+1
                print("Starting index:",starting) #Still need to generate index
                
    else:
        starting = number_of_index_done[0].iloc[-1]+1
        print("Starting index:",starting) #Still need to generate index

else:
    logging.info("no file exist, generate")
    wb = Workbook()
    ws = wb.active
    ws.title = excel_sheet_name
    wb.save(filename = file_name)
    starting = 0
    
# manually generate the different random seed
def random_generater():
    return(random.randint(1, 100000))

# Model Complexity (NN)
def nn_model_complexity_multiprocessing(X_train, X_val, y_train, y_val, X_test,best_params):
    #Missing data imputatoin 
    X_train = scaler_preprocessor.fit_transform(X_train)
    X_val = scaler_preprocessor.transform(X_val)
    X_test = scaler_preprocessor.transform(X_test)
    
    model = tf.keras.models.Sequential()
    # First hidden layer with input shape
    model.add(tf.keras.layers.Dense(best_params['hidden_layer_sizes'][0], input_shape=(X_train.shape[1],), activation=best_params['activation']))   
    for i in range(1,len(best_params['hidden_layer_sizes'])):
        # from second hidden layer to number of hidden layers
        model.add(tf.keras.layers.Dense(best_params['hidden_layer_sizes'][i], activation=best_params['activation']))
    # Ouput layer 
    model.add(tf.keras.layers.Dense(1,activation='sigmoid'))
    model.compile(loss='binary_crossentropy', optimizer=Adam(learning_rate=best_params['learnRate']), metrics=['accuracy'])   
    es = EarlyStopping(monitor='val_loss',mode='min', verbose=0, patience=30)     
    model.fit(X_train, y_train, validation_data=(X_val, y_val), verbose=0, 
                       batch_size = best_params['batch_size'], epochs = 100, 
                       callbacks=[es])
    
    difficulty = model.predict(X_test, verbose=0)
    #print(difficulty[0][0])
    return(difficulty[0][0])
    
if __name__ == "__main__":
    row_values = []
    start_time_overall = datetime.now()
    for index in range(starting,len(telco_X)):
        logging.info('running index:%s'%index)
        start_time_one = datetime.now()
        
        X_overall = telco_X
        y_overall = telco_y

        X_test = X_overall.iloc[[index]] # the test case that want to check the difficulty
        y_test = y_overall[index]

        X_without_test = X_overall.drop(index=[index]) # X,y the dataset wilthout the test case
        y_without_test = y_overall.drop(index=[index])
    
        #Call hyperparameter
        param = hyper_param_df.iloc[index]

        X_train_dataset = []
        X_val_dataset = []
        y_train_dataset = []
        y_val_dataset = []
        for nn in range(number_of_iter):
            X_train, X_val, y_train, y_val = train_test_split(X_without_test, y_without_test, test_size=0.3, random_state=random_generater())
            X_train_dataset.append(X_train)
            X_val_dataset.append(X_val)
            y_train_dataset.append(y_train)
            y_val_dataset.append(y_val)
       
    
        predicted_probabilities = []      
        def collect_result(result):
            print(result)
            predicted_probabilities.append(result)
            
        # Create a list of argument tuples
        arg_list = [(X_train_dataset[mm], X_val_dataset[mm], y_train_dataset[mm], y_val_dataset[mm]
                     ,X_test,param) for mm in range(number_of_iter)]
      
        try:
            pool = Pool(processes=number_of_cpu) 
            for args in arg_list:
                pool.apply_async(nn_model_complexity_multiprocessing, args, callback=collect_result)
            pool.close()
            pool.join()
                
        except KeyboardInterrupt:
            print("KeyboardInterrupt detected. Terminating...")
            predicted_probabilities = []
            pool.terminate()
            pool.join()

        end_time_one = datetime.now()
        one_case_time = end_time_one-start_time_one
        logging.info('one case_running_time:%s'%one_case_time)

        if len(predicted_probabilities) != number_of_iter:
            sys.exit("Error: Number of predicted_probabilities does not match with number_of_iter.")
        
        else:
            #index, X_test, y_test, predicted_probabilities(number: number_of_iter)
            result_array = np.concatenate([arr.flatten() for arr in predicted_probabilities])
            result_array = result_array.tolist()
            row_values = [index] + [X_test[column][index] for column in X_test.columns] + [y_test]+ result_array 
            
            wb = load_workbook(filename=file_name)
            ws = wb[excel_sheet_name]
            ws.append(row_values)
            wb.save(file_name)

end_time_overall = datetime.now()
overall_case_time = end_time_overall-start_time_overall
logging.info('overall_running_time:%s'%overall_case_time)
logging.info('End of the prediction_100:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
