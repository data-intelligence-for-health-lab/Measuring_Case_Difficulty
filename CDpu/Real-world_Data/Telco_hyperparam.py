import os
os.environ["CUDA_VISIBLE_DEVICES"]="-1"   
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

# scikit-learn k-fold cross-validation
from numpy import array
from sklearn.model_selection import KFold
import pandas as pd
import numpy as np

import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
from tensorflow.keras.optimizers import Adam

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, OneHotEncoder, StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer

import itertools
import random
from keras.callbacks import EarlyStopping
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

from ray import tune
from ray.tune.search.hyperopt import HyperOptSearch
import ray
import logging

ray.init(num_cpus=50, num_gpus=0)

log_file_name = './app3_telco_hyperparam.log'
file_name = './app3_telco_hyperparam.xlsx'
excel_sheet_name = '(telco)'

# Configure the logger
logging.basicConfig(filename=os.path.join(log_file_name), level=logging.DEBUG)
logging.info('Start of the app3:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

# Load Data
data = pd.read_csv('./WA_Fn-UseC_-Telco-Customer-Churn.csv') 
df = data

# Drop Customer ID
df = df.drop('customerID',axis=1)
# Convert SeniorCitizen Column to the object
df["SeniorCitizen"] = df["SeniorCitizen"].replace({0:"No",1:"Yes"})
# Convert Churn column to a numeric 0/1 variable
df["Churn"] = df["Churn"].replace({"No":0, "Yes":1})

# Missing data imputation use the same row information
df.loc[df['tenure'] == 0, 'TotalCharges'] = 0
df["TotalCharges"] = df["TotalCharges"].astype('float64')

# Categorical columns and numeric columns
Cat_col = ['gender', 'SeniorCitizen', 'Partner', 'Dependents', 'PhoneService',
       'MultipleLines', 'InternetService', 'OnlineSecurity', 'OnlineBackup',
       'DeviceProtection', 'TechSupport', 'StreamingTV', 'StreamingMovies',
       'Contract', 'PaperlessBilling', 'PaymentMethod']
Num_col = ["tenure", "TotalCharges", "MonthlyCharges"]


# Label Encoder for categorical 
label_encoder = LabelEncoder()
df[Cat_col] = df[Cat_col].apply(label_encoder.fit_transform)


# X and y set
X = df.drop('Churn',axis=1)
y = df['Churn']


# One-hot encoding for categorical 
onehot_encoder = OneHotEncoder(handle_unknown='ignore') 
enc_data = pd.DataFrame(onehot_encoder.fit_transform(X[Cat_col]).toarray())
enc_data.columns = onehot_encoder.get_feature_names_out()
column_names = enc_data.columns.to_list()
names = Num_col + column_names


# Merge numeric columns and one-hot encoded columns
X.drop(X[Cat_col] ,axis=1, inplace=True)
X_OH_data = pd.concat([X, enc_data], axis=1)
X_OH_data


# Scaler Set
scaler_transformer = Pipeline(steps=[('scaler', StandardScaler())])
scaler_preprocessor = ColumnTransformer(transformers=[
    ('scaling', scaler_transformer, Num_col)], remainder='passthrough')


#Split X and y
telco_X = X_OH_data
telco_y = y


# reload file
if os.path.isfile(file_name):
    wb = load_workbook(file_name, read_only=True)   # open an Excel file and return a workbook
    if excel_sheet_name in wb.sheetnames:
        logging.info('sheet exist:%s'%excel_sheet_name)
        number_of_index_done = pd.read_excel(file_name,sheet_name=excel_sheet_name,header=None)
        #print(number_of_index_done)

        if len(number_of_index_done) == 0:
            starting = len(number_of_index_done)
            print("Starting index:",starting) #when the excel file is empty

        else:
            if len(number_of_index_done) == len(data):
                print('Done')
                starting = None
            
            else:
                starting = number_of_index_done[0].iloc[-1]+1
                print("Starting index:",starting) #Still need to generate index
                
    else:
        starting = number_of_index_done[0].iloc[-1]+1
        print("Starting index:",starting) #Still need to generate index

else:
    logging.info("no file exist, generate")
    wb = Workbook()
    ws = wb.active
    ws.title = excel_sheet_name
    wb.save(filename = file_name)
    starting = 0
    
    
numbers = [5, 10, 15, 20]
one_combinations_with_order = list(itertools.product(numbers))
two_combinations_with_order = list(itertools.product(numbers, repeat=2))
three_combinations = list(itertools.product(numbers, repeat=3))

layer_neuron_orders = []
layer_neuron_orders = one_combinations_with_order + two_combinations_with_order+three_combinations

# manually generate the different random seed
def random_generater():
    return(random.randint(1, 100000))

def objective(config):
    X_train, X_val, y_train, y_val = train_test_split(X_without_test, y_without_test, test_size=0.3, random_state=random_generater())
    X_train_scaled = scaler_preprocessor.fit_transform(X_train)
    X_val_scaled = scaler_preprocessor.transform(X_val)
    
    # Create model
    model = tf.keras.models.Sequential()
    # First hidden layer with input shape
    model.add(Dense(config['hidden_layer_sizes'][0], input_shape=(X_train.shape[1],), activation=config['activation']))   
    for i in range(1,len(config['hidden_layer_sizes'])):
        # from second hidden layer to number of hidden layers
        model.add(Dense(config['hidden_layer_sizes'][i], activation=config['activation']))
    model.add(tf.keras.layers.Dense(1,activation='sigmoid'))
    model.compile(loss='binary_crossentropy', optimizer=Adam(learning_rate=config['learnRate']), metrics=['accuracy'])  
    
    es = EarlyStopping(monitor='val_loss',mode='min', verbose=0, patience=30)     
    result = model.fit(X_train_scaled, y_train, validation_data=(X_val_scaled, y_val), verbose=0, 
                       batch_size = config['batch_size'], epochs = 100, 
                       callbacks=[es])
    
    val_loss = result.history['val_loss'][-1]
    return {"val_loss": val_loss}



start_time_overall = datetime.now()
for index in range(starting,len(telco_X)):
    start_time_one = datetime.now()
    X_overall = telco_X
    y_overall = telco_y

    X_test = X_overall.iloc[[index]] # the test case that want to check the difficulty
    y_test = y_overall[index]

    X_without_test = X_overall.drop(index=[index]) # X,y the dataset wilthout the test case
    y_without_test = y_overall.drop(index=[index])
    
    search_space = {'learnRate': tune.choice([0.01,0.03,0.1]),
                    'batch_size': tune.choice([32,64,128]),
                    'activation':tune.choice(['relu','tanh']),
                    'hidden_layer_sizes':tune.choice(layer_neuron_orders)}

    algo = HyperOptSearch()

    tuner = tune.Tuner(
        objective,
        tune_config=tune.TuneConfig(
            metric="val_loss",
            mode="min",
            num_samples = 100,
            search_alg=algo,
        ),
        run_config = ray.air.config.RunConfig(verbose=0),
        param_space=search_space,
    )

    try:
        results = tuner.fit()
        
        print("Best hyperparameters found were: ", results.get_best_result().config)
        hyper_param = results.get_best_result().config

        rows = [index,hyper_param['learnRate'],hyper_param['batch_size'],
                hyper_param['activation'],hyper_param['hidden_layer_sizes']] 
        logging.info('rows:%s'%rows)
        row_with_string_tuple = rows[:-1] + [str(rows[-1])]

        wb = load_workbook(filename=file_name)
        ws = wb[excel_sheet_name]
        ws.append(row_with_string_tuple)
        wb.save(file_name)

        end_time_one = datetime.now()
        one_case_time = end_time_one-start_time_one
        logging.info('one case_running_time:%s'%one_case_time)


    except KeyboardInterrupt:
        results = None
        ray.shutdown()
        break  

end_time_overall = datetime.now()    
overall_case_time = end_time_overall-start_time_overall
logging.info('overall_running_time:%s'%overall_case_time)
logging.info('End of the app3:' + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
